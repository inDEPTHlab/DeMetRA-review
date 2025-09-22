# /// script
# requires-python = ">=3.13"
# dependencies = [
#     "openpyxl==3.1.5",
# ]
# ///

import marimo

__generated_with = "0.15.0"
app = marimo.App(width="medium")


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    # DeMetRA - literature review

    This the preprocessing pipeline for the metadata included in the DeMetRA review. Inputs to the pipeline are: 

    - The `MPS_review_systematic_DATE.xlsx` file, which contains the metadata manually extracted by Isabel 
    - The bibliography file, to supplement more information about the publications included in the review.
    """
    )
    return


@app.cell
def _():
    import marimo as mo

    import pandas as pd
    import openpyxl
    import numpy as np

    from dateutil.parser import parse as parse_date

    assets_directory = './assets/'
    return assets_directory, mo, parse_date, pd


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    In the file created by Isabel, there are 3 sheets:

    1. The main list of MPSs / papers included in the review
    2. The base sample references papers for summary statistics 
    3. The base sample references papers for validated algorithms
    """
    )
    return


@app.cell
def _(assets_directory, mo, pd):
    sys_review_file = 'MPS_review_systematic_2025-08-14.xlsx' # 'MPS_review_systematic_2025-02-14.xlsx'

    lit_raw, base_ss, base_va = pd.read_excel(f'{assets_directory}/Annotations_raw_files/{sys_review_file}',
                                              sheet_name=[0, 1, 2]).values()

    lit_raw = lit_raw.loc[lit_raw.Include == 'Yes'].drop('Include', axis=1) # All included

    # Drop any empty rows 
    base_ss = base_ss.loc[base_ss.Identifier.notna() & (base_ss.Identifier.str.strip() != "")]
    base_va = base_va.loc[base_va.Identifier.notna() & (base_va.Identifier.str.strip() != "")]

    if 'Base use' not in base_ss.columns:
        base_ss.insert(1, 'Base use', 'Summary statistics')

    if 'Base use' not in base_va.columns:
        base_va.insert(1, 'Base use', 'Validated algorithm')

    # print(base_ss.shape, base_va.shape)

    print(lit_raw.shape, '\n', list(lit_raw.columns))
    print(base_ss.shape, '\n', list(base_ss.columns))
    print(base_va.shape, '\n', list(base_va.columns))

    lit_base_raw = pd.concat([base_ss, base_va], axis=0, ignore_index=True, sort=False)

    df_tabs = mo.ui.tabs({"Main": lit_raw, "Base": lit_base_raw})
    return df_tabs, lit_base_raw, lit_raw


@app.cell
def _(df_tabs):
    df_tabs
    return


@app.cell
def _(mo):
    mo.md(r"""### General helper functions""")
    return


@app.cell
def _(lit_base_raw, lit_raw, mo, pd):
    def check_variable_levels(var):
        allobs = pd.concat([lit_raw[var], lit_base_raw[var]])

        counts = allobs.value_counts(dropna=False)
        counts.index = [f"~{i}~" for i in counts.index]

        var_multiple = f'Multiple_{var.lower()}'

        try: 
            allobs_multiple = pd.concat([lit_raw[var_multiple], lit_base_raw[var_multiple]])
            counts_multiple = allobs_multiple.value_counts(dropna=False)
            counts_multiple.index = [f"~{i}~" for i in counts_multiple.index]

            return counts, counts_multiple, # pd.concat([allobs, allobs_multiple], axis=1)

        except:

            return counts, mo.md(f"No {var_multiple} found.")


    def inspect_variable(variable, variable_base = None):

        if not variable_base: 
            variable_base = variable

        df = pd.concat([lit_raw[variable], lit_base_raw[variable_base]])
        counts = df.value_counts(dropna=False).sort_index()

        return(counts)


    def count_categories_per_phenotype(data):
        '''Check that no phenotypes are assigned to multiple categories '''
        counts = data.groupby('Phenotype')['Category'].nunique().sort_values(ascending=False)
        if counts.iloc[0] > 1: 
            return counts
        else:
            return "::ok::"


    def aggregate_values(series):
        if series.apply(lambda x: isinstance(x, list)).sum() > 0:
            return series.iloc[0]
        unique_values = series.unique()
        if len(unique_values) == 1:
            return unique_values[0]
        else:
            return list(unique_values)


    def replace_multiples(df, drop_multiple_cols=True):
        for var in ['Tissue', 'Array', 'Ancestry']:
            if var == 'Ancestry':
                df.loc[df[var] == 'Multiple', var] = 'Mixed'
            else:
                df.loc[df[var] == 'Multiple', var] = [f'Multiple ({values})' for values in df.loc[
                    df[var] == 'Multiple', f'Multiple_{var.lower()}']]

            if drop_multiple_cols:
                df = df.drop([f'Multiple_{var.lower()}'], axis=1)
        return df


    def coerce_to_numeric(df, old_var, new_var):

        df[new_var] = pd.to_numeric(df[old_var], errors='coerce')

        coerced_mask = df[new_var].isna() & df[old_var].notna()
        coerced_values = df.loc[coerced_mask, old_var].value_counts()

        if coerced_values.shape[0] > 0:
            print('\nThese values are coerced to NA:\n', coerced_values)

        return df


    def clean_n_CpGs(df, replace_with_df = None):

        n_cpg_var = "Number of CpGs"
        new_n_cpg_var = "n CpGs"

        empty_mps_count = df[df[n_cpg_var] == 0].shape[0]
        if empty_mps_count > 0:
            print(f"Dropping {empty_mps_count} MPSs which did no identify a solution (0 CpGs included)")
            df = df[df[n_cpg_var] != 0].reset_index(drop=True)

        na_count = df[n_cpg_var].isna().sum()
        print(f"{na_count} NA values in `{n_cpg_var}`")

        if replace_with_df is not None:
            print("    --> Replacing with info from base, where possible...")

            for _idx, _row in df.iterrows():
                if pd.isna(_row[n_cpg_var]) & (_row['What_is_available'] != "Only phenotype"):
                    match = replace_with_df[(replace_with_df['Identifier'] == _row['Identifier_base']) & 
                                            (replace_with_df['Phenotype'] == _row['Phenotype'])]
                    if match.empty:
                        print(f"\t({_row['What_is_available']}) {_row['Identifier_base']} {_row['Phenotype']} not found.")
                    else:
                        df.at[_idx, n_cpg_var] = match.iloc[0][n_cpg_var]

            print(f"{df[n_cpg_var].isna().sum()} `NA` values left +",
                  f"{df.loc[df[n_cpg_var] == 'Not reported',].shape[0]} `Not reported` values.")

        df = coerce_to_numeric(df, n_cpg_var, new_n_cpg_var)

        return df

    ris_tags = {'TY': 'Reference Type', 
                'AU': 'Author_list', 
                'PY': 'Year', 
                'TI': 'Title', 
                'T2': 'Journal', 
                'J2': 'Journal', 
                'AB': 'Abstract', 
                'DO': 'DOI', 
                'UR': 'URL', 
                'KW': 'Keywords', 
                'DA': 'Date'}

    def parse_ris(file_path):
        with open(file_path, 'r', encoding='utf-8') as _file:
            references = []
            current_entry = {}
            current_tag = None
            for line in _file:
                if not line.strip():
                    continue
                elif line.strip() == 'ER  -':
                    references.append(current_entry)
                    current_entry = {}
                    current_tag = None
                else:
                    tag = line[:5]
                    if tag in [f'{_t}  -' for _t in ris_tags.keys()]:
                        current_tag = tag
                        value = line[5:].strip()
                        key = ris_tags[tag[:2]]
                        if key in current_entry:
                            if isinstance(current_entry[key], list):
                                current_entry[key].append(value)
                            else:
                                current_entry[key] = [current_entry[key], value]
                        else:
                            current_entry[key] = value
                    elif tag in [f'{y}-' for y in range(2000, 2025)]:
                        value = line.strip()
                        if 'Date' not in current_entry:
                            current_entry['Date'] = value
                    elif current_tag == 'KW  -':
                        value = line.strip()
                        if isinstance(current_entry['Keywords'], list):
                            current_entry['Keywords'].append(value)
                        else:
                            current_entry['Keywords'] = [current_entry['Keywords'], value]
            return references
    return (
        aggregate_values,
        check_variable_levels,
        clean_n_CpGs,
        coerce_to_numeric,
        count_categories_per_phenotype,
        inspect_variable,
        parse_ris,
        replace_multiples,
    )


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ## STEP 1 - Data cleaning
    ### Data inspection

    To guide manual cleaning of the excel file, I print value counts for the main columns of interest.

    - 'Tissue', 'Array', 'Ancestry', 'Developmental_period'
    - 'Determining_weights_*'
    - 'Including_CpGs_*'
    """
    )
    return


@app.cell
def _(check_variable_levels, mo):
    # Inspect values 
    # NOTE: this is to guide manual cleaning of the excel file!

    vars_to_inspect = ['Tissue', 'Array', 'Ancestry', 'Developmental_period']

    mo.ui.tabs({v: check_variable_levels(v) for v in vars_to_inspect})
    return


@app.cell(hide_code=True)
def _(lit_base_raw, lit_raw, mo, pd):
    def summarize_dimension_reduction_strategies():

        strategies = [f"Including_CpGs_{i}" for i in range(1, 6)] # five levels

        df = pd.concat([lit_raw[strategies], lit_base_raw[strategies]])

        tab_dict = {s: df[s].value_counts(dropna=False).sort_index() for s in strategies}

        stacked_strategies = df.stack().reset_index(drop=True)
        conbined_strategies = df.astype(str).agg(" --- ".join, axis=1)

        tab_dict["All_strategies"] = stacked_strategies.value_counts(dropna=False).sort_index()
        tab_dict["Combined_strategies"] = conbined_strategies.value_counts(dropna=False)

        tabs = mo.ui.tabs(tab_dict)
        return(tabs)

    summarize_dimension_reduction_strategies()
    return


@app.cell
def _(inspect_variable):
    # Summarize weight estimation strategies
    inspect_variable("Determining_weights_1")
    return


@app.cell
def _(inspect_variable):
    # Summarize internal validation strategies
    inspect_variable("Train_test", "Train_validate")
    return


@app.cell
def _(inspect_variable):
    # Summarize external validation strategies
    inspect_variable("Independent_validation", "Independent_test")
    return


@app.cell
def _(inspect_variable):
    # Summarize performance 
    inspect_variable("Reflect_phenotype")
    return


@app.cell
def _(inspect_variable):
    inspect_variable("Sample_overlap_target_base")
    return


@app.cell
def _(count_categories_per_phenotype, lit_base_raw, lit_raw):
    # Check that no phenotypes are assigned to multiple categories 
    count_categories_per_phenotype(lit_raw)
    count_categories_per_phenotype(lit_base_raw)
    return


@app.cell
def _():
    # Covariates
    # Sample_overlap_target_base
    # 'Comparison', 'Missing_value_note', 'Link'
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ###  Bibliography file
    I additionally parse the bibliography file for publication dates, abstracts and keywords.

    <ins>Note</ins>: the `Bibliography_<date>.txt` files are generated by exporting the`.ris` or `.enl` files as a RefMan (RIS) .txt file (using Endnote).
    """
    )
    return


@app.cell
def _(assets_directory, lit_raw, parse_date, parse_ris, pd):
    parsed_data = parse_ris(f"{assets_directory}/Bibliography_raw_files/Bibliography_2025-07-25.txt")

    bib = pd.DataFrame(parsed_data)[['Author_list', 'Year', 'Title', 'Journal', 'Keywords', 'Abstract', 'Date', 'DOI']] # 'URL' no reliable parsing, use links in main 
    print(bib.shape)

    bib_incl = bib.loc[bib.Title.isin(lit_raw.Title.unique()),].reset_index(drop=True)
    print(bib_incl.shape)

    # Cleaning Dates
    print('Note', bib_incl.Date.isna().sum(), 'NaN Date values will be set to 01/06 of respective year')
    date_tmp = pd.Series([' '.join([d, y]) if y not in d else d for d, y in zip(bib_incl.Date.map(str), bib_incl.Year.map(str))])
    date_tmp[date_tmp=="May-Jun 2025"] = "Jun 2025"
    bib_incl.loc[:, 'Date'] = date_tmp.apply(lambda date: 
                                             parse_date(date).strftime('%Y-%m-%d') if 'nan' not in date else 
                                             parse_date('01 06' + date[3:]).strftime('%Y-%m-%d'))

    # Cleaning Authors
    bib_incl['Author'] = [f"{fa[0].split(',')[0]} et al." if len(fa) > 1 else f"{fa[0].split(',')[0]}" 
                          for fa in bib_incl['Author_list']]

    # Cleaning Journal names
    bib_incl.loc[bib_incl.DOI.str.contains('10.1007/s00787-024-02390-1', na=False), 'Journal'] = 'Eur Child Adolesc Psych'
    bib_incl.loc[bib_incl.DOI.str.contains('10.3390/ijms', na=False),  'Journal'] = 'Int J Mol Sci'
    bib_incl.loc[bib_incl.DOI.str.contains('10.3390/genes16050591', na=False),  'Journal'] = 'Genes'

    bib_incl
    return (bib_incl,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    ### Data cleaning 

    1. Extract and remove the **`"*_multiple"`** `Tissue`, `Array`, and `Ancestry` columns so that only one column is kept in the dataset.
       I do this for both the main list and the base samples.

    3. Clean the **`number of CpGs`** column by replacing `NA` with base sample sizes and setting "Not reported" to NA.
       I also remove 2 scores that have `Number of CpGs == 0` (because they did not identify a solution).

    5. Clean the **"Sample size"** columns so that they are all numeric.
    """
    )
    return


@app.cell
def _(
    bib_incl,
    clean_n_CpGs,
    coerce_to_numeric,
    lit_base_raw,
    lit_raw,
    replace_multiples,
):
    # Clean *_multiple variables ------------------------------------------------------------
    lit_main = replace_multiples(lit_raw)
    lit_base = replace_multiples(lit_base_raw)

    # Clean Number of CpGs ------------------------------------------------------------------

    lit_base = clean_n_CpGs(lit_base)
    # Note: if this is missing, I attempt retrieving the value form base whenever possible, 
    # but these are assumed to match (checked later)
    lit_main = clean_n_CpGs(lit_main, replace_with_df = lit_base)

    # Clean Sample size columns -------------------------------------------------------------

    lit_main = coerce_to_numeric(lit_main, 'Sample_size_total', 'Sample size')
    lit_main = coerce_to_numeric(lit_main, 'Sample_size_case', 'n Cases')
    lit_main = coerce_to_numeric(lit_main, 'Sample_size_control', 'n Controls')

    lit_base = coerce_to_numeric(lit_base, 'Sample_size_total', 'Sample size')
    lit_base = coerce_to_numeric(lit_base, 'Sample_size_case', 'n Cases')
    lit_base = coerce_to_numeric(lit_base, 'Sample_size_control', 'n Controls')


    # lit_main = lit_main.rename(columns={'Author': 'Author_dirty', 
    #                                     'Journal': 'Journal_dirty'}).merge(bib_incl, on='Title', how='left', suffixes=['', '_BIB'])
    lit_main = lit_main.rename(columns={'Author': 'Author_dirty'}).merge(bib_incl, on='Title', how='left', suffixes=['', '_BIB'])

    # lit_main['Journal'] = lit_main['Journal'].apply(lambda x: x[0] if isinstance(x, list) else x)
    # lit_main['DOI_BIB'] = lit_main['DOI_BIB'].fillna(lit_main['DOI'])
    # lit_main['DOI'] = lit_main['DOI_BIB']
    # lit_main.drop(['DOI_BIB', 'Year_BIB'], axis=1, inplace=True)
    return lit_base, lit_main


@app.cell
def _(lit_main):
    lit_main.loc[lit_main.Journal != lit_main.Journal_BIB, ['Journal', 'Journal_BIB','Title']]
    # lit_main.loc[lit_main.Author != lit_main.Author_BIB, ['Author', 'Author_BIB','Title']]
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""### Rename and reorder""")
    return


@app.cell
def _(lit_base, lit_main):
    # lit_main.What_is_available.value_counts(dropna=False)

    rename_based_on = {
        "Only phenotype": "Raw individual-level data",
        "Validated MPS algorithm": "Pre-established MPS",
        "EWAS summary statistics": "Published summary statistics (semi-supervised)"
    }

    lit_main["Based on"] = lit_main["What_is_available"].map(rename_based_on) # .fillna(df["What_is_available"]) # no NA values here


    rename_cols_dict = {'Sample_type': 'Sample type', 
                        'Developmental_period': 'Developmental period', 
                        'Including_CpGs_1' : 'Dimension reduction (1)', 
                        'Including_CpGs_2': 'Dimension reduction (2)', 
                        'Including_CpGs_3': 'Dimension reduction (3)',  
                        'Including_CpGs_4': 'Dimension reduction (4)',  
                        'Including_CpGs_5': 'Dimension reduction (5)', 
                        'Determining_weights_1': 'Weights estimation',
                        'Train_test': 'Internal validation',
                        'Train_validate': 'Internal validation',
                        'Independent_validation': 'External validation',
                        'Independent_test': 'External validation',
                        'Reflect_phenotype': 'Performance',
                        'Type': 'Publication type', 
                        'Link': 'URL'}

    mps_table = lit_main.rename(columns=rename_cols_dict)
    base_table = lit_base.rename(columns=rename_cols_dict)

    return base_table, mps_table


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    ### Publications table 
    Create a  table grouped by "paper" rather than by "MPS"
    """
    )
    return


@app.cell
def _(aggregate_values, mps_table):
    pub_table = mps_table.groupby('Title').agg(aggregate_values).reset_index()
    n_MPS = mps_table.Title.value_counts()
    for _t in n_MPS.index:
        pub_table.loc[pub_table.Title == _t, 'n MPSs'] = int(n_MPS[_t])
    return (pub_table,)


@app.cell
def _(mo):
    mo.md(
        r"""
    ### Mearge target and base samples

    <ins>Note</ins>:&emsp;base -> "development dataset"; <br>&emsp;&emsp;&emsp; target -> "application dataset"
    """
    )
    return


@app.cell
def _(base_table, mps_table):
    targ_only = mps_table.loc[mps_table['Based on'] != 'Raw individual-level data',]
    print(targ_only.shape, base_table.shape)
    return (targ_only,)


@app.cell
def _(targ_only):
    multi_base_id = targ_only.loc[targ_only.Identifier_base == 'Multiple',]
    multi_base_pubs = multi_base_id.Title.unique()
    print(f"{multi_base_id.shape[0]} MPSs (from {len(multi_base_pubs)} publications) have multiple base samples.")
    print(multi_base_id['Based on'].value_counts())
    print("\n", targ_only['Multiple_identifier_base'].value_counts())
    return


@app.cell
def _(base_table, targ_only):
    # For Multiple base IDs linked to the same target MPS: split identifiers into rows
    targ_only.loc[:, "Identifier_base_expanded"] = targ_only.apply(
        lambda row: [row["Identifier_base"]] if row["Identifier_base"] != "Multiple" 
        else row["Multiple_identifier_base"].split("; "), axis=1
    )

    # And unpack multiples
    targ_expanded = targ_only.explode("Identifier_base_expanded")

    # For base IDs with multiple rows (i.e., phenotypes), append phenotype info to the IDs to they are truly unique
    base_id_counts = base_table.groupby("Identifier")["Identifier"].transform("count")
    base_id_mask = base_id_counts > 1
    targ_id_mask = targ_expanded["Identifier_base_expanded"].isin(set(base_table.loc[base_id_mask, "Identifier"]))

    # If an Identifier appears more than once → append Phenotype
    base_table.loc[:, "Identifier_base_expanded"] = base_table["Identifier"]
    base_table.loc[base_id_mask, "Identifier_base_expanded"] = (base_table["Identifier"] + base_table["Phenotype"])

    targ_expanded.loc[targ_id_mask, "Identifier_base_expanded"] = (
        targ_expanded.loc[targ_id_mask, "Identifier_base_expanded"] + targ_expanded.loc[targ_id_mask, "Phenotype"])

    # In case the same base paper is used for summary stats and validated score, add info to IDs to ensure they are still unique
    base_table_use = ['ss' if use == 'Summary statistics' else 'va' for use in base_table["Base use"]]
    targ_table_use = ['ss' if use == 'Published summary statistics (semi-supervised)' else 'va' for use in targ_expanded['Based on']]

    base_table.loc[:, "Identifier_base_expanded"] = base_table["Identifier_base_expanded"] + base_table_use
    targ_expanded.loc[:, "Identifier_base_expanded"] = targ_expanded["Identifier_base_expanded"] + targ_table_use
    return (targ_expanded,)


@app.cell
def _(base_table):
    count_duplicate_base_ids = base_table['Identifier_base_expanded'].value_counts()

    duplicate_base_ids = list(count_duplicate_base_ids.index[count_duplicate_base_ids > 1])
    base_table.loc[base_table['Identifier_base_expanded'].isin(duplicate_base_ids), :]

    # TMP: I pick duplicate with more info, drop the others
    duplicate_idx = [155, 156]
    base_table_tmp = base_table[~base_table.index.isin(duplicate_idx)]
    return (base_table_tmp,)


@app.cell
def _(base_table_tmp, targ_expanded):
    # Check mismatches
    targ_ids = set(targ_expanded['Identifier_base_expanded'])
    base_ids = set(base_table_tmp['Identifier_base_expanded'])

    common_ids = targ_ids & base_ids
    unique_to_base = base_ids - targ_ids
    unique_to_targ = targ_ids - base_ids

    print(len(common_ids), 'matching ids\n')

    print(f"{len(unique_to_base)} / {len(base_ids)} unique to base")
    if (len(unique_to_base) > 0): [print('\t', p) for p in unique_to_base]

    print(f"{len(unique_to_targ)} / {len(targ_ids)} unique to targ")
    if (len(unique_to_targ) > 0): [print('\t', p) for p in unique_to_targ]

    # Remove "Unclear" base links
    print('\nCleaning...\n')
    targ_expanded_clean = targ_expanded.loc[targ_expanded["Identifier_base_expanded"] != "Unclearva", :]
    targ_ids = set(targ_expanded_clean['Identifier_base_expanded'])
    unique_to_targ = targ_ids - base_ids
    print(f"{len(unique_to_targ)} / {len(targ_ids)} unique to targ")
    if (len(unique_to_targ) > 0): [print('\t', p) for p in unique_to_targ]

    # Inspect 
    # mo.ui.tabs({"Main": targ_expanded.loc[targ_expanded['Identifier_base_expanded'].isin(unique_to_targ), :], 
    #             "Base": d_base.loc[d_base['Identifier_base_expanded'].isin(unique_to_base), :], 
    #            })
    return (targ_expanded_clean,)


@app.cell
def _(aggregate_values, base_table_tmp, targ_expanded_clean):
    # Now I can merge with base 
    targ_base_merge = targ_expanded_clean.merge(base_table_tmp, on="Identifier_base_expanded", how="inner", # "left", 
                                            suffixes=[' [application]', ' [development]'])

    # Group back by original targ row
    targ_base_aggre = targ_base_merge.groupby(targ_expanded_clean.index).agg(aggregate_values).reset_index(drop=True)

    print(targ_base_aggre.shape, sorted(targ_base_aggre.columns))
    targ_base_aggre
    # d_targ_base.to_csv(f'{assets_directory}MPS_base_target_cleaned.csv', index=False)
    return targ_base_aggre, targ_base_merge


@app.cell
def _(targ_base_merge, targ_expanded_clean):
    print(targ_expanded_clean.shape, targ_base_merge.shape)
    return


@app.cell
def _(mo, mps_base_matched):
    def detect_mismatch(variable, df=mps_base_matched, x=' [application]', y=' [development]'):
        df_subset = df.loc[df[f"{variable}{x}"] != df[f"{variable}{y}"], :]
        value_counts = (df_subset[f"{variable}{x}"].astype(str) +" --- "+ 
                        df_subset[f"{variable}{y}"].astype(str)).value_counts()
        return df_subset, value_counts

    vars_to_inspect_match = ['Phenotype', 'Category', 'Developmental period', 'Tissue', 'Array', 'Ancestry',
                       # Dimension reduction (n)
                       # Weights estimation
                       # Internal / external validation
                       # Performance
                      'Sample type']

    # diff: "n CpGs"

    mo.ui.tabs({v: detect_mismatch(v) for v in vars_to_inspect_match})
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""###Save cleaned files""")
    return


@app.cell
def _(mo, mps_table, pub_table, targ_base_aggre):
    # Subset and reorder
    mps_table_clean = mps_table[['Phenotype', 'Category', 'Author', 'Year', 'Title', 'DOI', 
                           'n CpGs', 'Based on', 'Sample size', 'n Cases', 'n Controls', 'Sample type', 
                           'Developmental period', 'Tissue', 'Array', 'Ancestry', 
                           'Publication type', 'Journal'] + 
                           ['Keywords', 'Abstract', 'Author_list', 'Date', 'URL'] + 
                           [f'Dimension reduction ({i})' for i in range(1, 6)] +
                           ['Weights estimation', 'Internal validation', 'External validation', 'Performance',
                            'Comparison', 'Missing_value_note', 'Covariates']]

    pub_table_clean = pub_table[['Author', 'Year', 'Title', 'DOI', 'Phenotype', 'Category', 'n MPSs',
                           'n CpGs', 'Based on', 'Sample size', 'n Cases', 'n Controls', 'Sample type', 
                           'Developmental period', 'Tissue', 'Array', 'Ancestry', 
                           'Publication type', 'Journal'] + 
                           ['Keywords', 'Abstract', 'Author_list', 'Date', 'URL'] + 
                           [f'Dimension reduction ({i})' for i in range(1, 6)] +
                           ['Weights estimation', 'Internal validation', 'External validation', 'Performance',
                            'Comparison', 'Missing_value_note', 'Covariates']]

    mps_base_matched = targ_base_aggre[['Phenotype [application]', 'Phenotype [development]', 
                                        'Category [application]', 'Category [development]', 
                                        'Author', 'Year', 'Title', 
                                        'n CpGs [application]', 'n CpGs [development]',
                                        'Based on', 
                                        'Sample size [application]', 'Sample size [development]', 
                                        'n Cases [application]', 'n Cases [development]', 
                                        'n Controls [application]', 'n Controls [development]', 
                                        'Sample type [application]', 'Sample type [development]', 
                                        'Developmental period [application]', 'Developmental period [development]', 
                                        'Tissue [application]', 'Tissue [development]', 
                                        'Array [application]', 'Array [development]', 
                                        'Ancestry [application]', 'Ancestry [development]', 
                                        'Dimension reduction (1) [application]', 'Dimension reduction (1) [development]', 
                                        'Dimension reduction (2) [application]', 'Dimension reduction (2) [development]', 
                                        'Dimension reduction (3) [application]', 'Dimension reduction (3) [development]', 
                                        'Dimension reduction (4) [application]', 'Dimension reduction (4) [development]', 
                                        'Dimension reduction (5) [application]', 'Dimension reduction (5) [development]', 
                                        'Weights estimation [application]', 'Weights estimation [development]', 
                                        'Internal validation [application]', 'Internal validation [development]', 
                                        'External validation [application]', 'External validation [development]',
                                        'Performance [application]', 'Performance [development]', 
                                        'Comparison [application]', 'Comparison [development]', 
                                        'Missing_value_note [application]', 'Missing_value_note [development]', 
                                        'Covariates [application]', 'Covariates [development]', 
                                        # 'URL [application]', 'URL [development]', 
                                        'Sample_overlap_target_base [application]', 'Sample_overlap_target_base [development]']]

    mo.ui.tabs({"Main": mps_table_clean, "Base-matched": mps_base_matched, "Pubs": pub_table_clean})
    return mps_base_matched, mps_table_clean, pub_table_clean


@app.cell
def _(assets_directory, mps_base_matched, mps_table_clean, pub_table_clean):
    mps_table_clean.to_csv(f'{assets_directory}MPS_table_cleaned.csv', index=False)
    pub_table_clean.to_csv(f'{assets_directory}Publication_table_cleaned.csv', index=False)

    mps_base_matched.to_csv(f'{assets_directory}MPS_base_matched_cleaned.csv', index=False)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""### Publication network graph""")
    return


@app.cell
def _():
    import ast
    import networkx as nx
    import pickle
    return ast, nx, pickle


@app.cell(disabled=True)
def _(ast, nx, pub_table_2):
    G = nx.Graph()
    for pub in pub_table_2.Title:
        title = f"Paper/{pub.replace(':', ' ')}"
        authors = ast.literal_eval(pub_table_2.loc[pub_table_2.Title == pub, 'Author_list'].iloc[0])
        for author in authors:
            author = f"Author/{author.replace('. ', '.')}"
            G.add_edge(author, title)
    return (G,)


@app.cell(disabled=True)
def _(G, nx):
    # Estimate optimal node positions ----------------------------------------------------
    # pos = nx.spring_layout(G, seed=3108, k=1000)
    pos = nx.fruchterman_reingold_layout(G, seed=3108, k=0.05)

    # Assign 'pos' attribute to the nodes in the graph
    for node in G.nodes:
        G.nodes[node]['pos'] = pos[node]

    # G.nodes
    return


@app.cell(disabled=True)
def _(G, assets_directory, pickle):
    with open(f'{assets_directory}/Publications_network.pkl', 'wb') as _file:
        pickle.dump(G, _file)
    return


@app.cell(disabled=True)
def _(assets_directory, pickle):
    with open(f'{assets_directory}/Publications_network.pkl', 'rb') as _file:
        G_1 = pickle.load(_file)
    return


if __name__ == "__main__":
    app.run()
